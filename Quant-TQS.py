# -*- coding: utf-8 -*-
"""
Created on Wed Feb  7 14:56:37 2024

@author: Mateus

Teste1
"""
import pandas as pd
import numpy as np

####
def preencher_nan(df):
    A = df.isnull().values

    out = df.values[np.argsort(A, axis=0, kind='mergesort'), np.arange(A.shape[1])]

    tt = pd.DataFrame(out, columns=df.columns)
    return tt
####

Tem_mais_prancha = False

# Abrir arquivo TABFER.txt, localizado na mesma pasta que o arquivo .py e criar
# uma lista com os valores separados por espaços e salvos como string
with open("TABFER.txt", "r") as fobj:
    all_lines = [[str(x) for x in line.split()] for line in fobj]

# Salvar lista criada como um DataFrame para ser utilizado como tabela    
dados = pd.DataFrame(data = all_lines)


# Retorna a quantidade de linha e colunas do DataFrame
linha_dados = dados.shape[0]
coluna_dados = dados.shape[1]

# Acha a linha i que a palavra "Planta" se encontra e salva a linha como c
for i, row in dados.iterrows():
    if row[0] == "Planta":
        c=i #Linha que possui a string "Planta" -- Título das pranchas
        break
    
# Salva a linha c como título
titulo = dados.iloc[[c]]
titulo = titulo.iloc[:,:6]
titulo.columns = ["Aço", "Bitola (mm)", "Comprimento (cm)", "Peso (kg)", " ", " "]

lista_titulos = titulo

for i, row in dados.iterrows():
    if row[0] == "RESUMO":
        RES1=i #Linha que possui a string "RESUMO" -- RESUMO DE AÇO
        break
DADOS_RES1 = RES1 + 4

dad1 = dados.iloc[DADOS_RES1:].reset_index(drop=True)


for i, row in dados.iterrows():
    if row[0] == "Resumo":
        RES2=i #Linha que possui a string "Resumo" -- Resumo de aço por elemento
        break    

# Cria DataFrame "dad2" equivalente ao quantitativo por planta
dad2 = dad1.iloc[:(RES2 - DADOS_RES1 - 2)].reset_index(drop=True)

dad2 = dad2.iloc[:,:6]

dad2.columns = ["Aço", "Bitola (mm)", "Comprimento (cm)", "Peso (kg)", " ", " "]

dad2_comb = pd.concat([titulo, dad2])

dad2_simp = dad2.pivot(columns="Bitola (mm)", values="Peso (kg)").reset_index(drop=True)

#Transforma string em float e ordena por bitola
d2 = dad2_simp.drop(["Total"], axis=1).transpose().reset_index().astype(float).sort_values(by="Bitola (mm)").transpose().dropna(how="all")
d2 = preencher_nan(d2)    
d2.columns = d2.iloc[0]
d2 = d2[1:].reset_index(drop=True).dropna(axis = 1, how="all").dropna(how="all")

d2_comb = d2

# Cria DataFrame "dad3" equivalente ao quantitativo por planta e por elemento
dad3 = dados.iloc[(c + 6):(RES1 - 1)].reset_index(drop=True)

dad3tt = dad3.dropna(how="all") #dad3 sem limitação de colunas

dad3 = dad3.iloc[:,:6].dropna(how="all")

dad3.columns = ["Aço", "Pos", "Bit (mm)", "Quant", "Comp unit (cm)", "Comp total (cm)"]

dad3_comb = dad3




# Cria DataFrame dadosP2, excluindo a Prancha já editada
dadosP2 = dados.iloc[(c+1):].reset_index(drop=True)

# Acha a linha i que a palavra "Planta" se encontra e salva a linha como c
for i, row in dadosP2.iterrows():
    if row[0] == "Planta":
        c=i
        Tem_mais_prancha = True
        break

while Tem_mais_prancha == True:
    # Salva a linha c como título
    titulo = dadosP2.iloc[[c]]
    titulo = titulo.iloc[:,:6]
    titulo.columns = ["Aço", "Bitola (mm)", "Comprimento (cm)", "Peso (kg)", " ", " "]
    
    lista_titulos = pd.concat([lista_titulos, titulo]).reset_index(drop=True)
    
    dadosP2_nr = dadosP2
    dadosP2 = dadosP2.iloc[(c+1):].reset_index(drop=True)
    
    for i, row in dadosP2.iterrows():
        if row[0] == "RESUMO":
            RES1=i
            break
    DADOS_RES1 = RES1 + 4
    
    for i, row in dadosP2_nr.iterrows():
        if row[0] == "RESUMO":
            RES1_nr=i
            break

    dad1 = dadosP2.iloc[DADOS_RES1:].reset_index(drop=True)

    for i, row in dadosP2.iterrows():
        if row[0] == "Resumo":
            RES2=i
            break    

    # Cria DataFrame "dad2" equivalente ao quantitativo por planta
    dad2 = dad1.iloc[:(RES2 - DADOS_RES1 - 2)].reset_index(drop=True)

    dad2 = dad2.iloc[:,:6]

    dad2.columns = ["Aço", "Bitola (mm)", "Comprimento (cm)", "Peso (kg)", " ", " "]
     
    dad2_comb = pd.concat([dad2_comb, titulo, dad2]).reset_index(drop=True)
    
    dad2_simp = dad2.pivot(columns="Bitola (mm)", values="Peso (kg)").reset_index(drop=True)
    
    #Transforma string em float 
    d2 = dad2_simp.drop(["Total"], axis=1).transpose().reset_index().astype(float).sort_values(by="Bitola (mm)").transpose().dropna(how="all")
    d2 = preencher_nan(d2)    
    d2.columns = d2.iloc[0]
    d2 = d2[1:].reset_index(drop=True).dropna(axis = 1, how="all").dropna(how="all")
    
    d2_comb = pd.concat([d2_comb, d2]).reset_index(drop=True) 
        
    # Cria DataFrame "dad3" equivalente ao quantitativo por planta e por elemento
    dad3 = dadosP2.iloc[(5):(RES1 - 1)].reset_index(drop=True)
    
    dad3t = dad3.dropna(how="all")

    dad3 = dad3.iloc[:,:6].dropna(how="all")

    dad3.columns = ["Aço", "Pos", "Bit (mm)", "Quant", "Comp unit (cm)", "Comp total (cm)"]
    
    dad3_comb = pd.concat([dad3_comb, dad3]).reset_index(drop=True)
    
    dad3tt = pd.concat([dad3tt, dad3t]).reset_index(drop=True) # dad3_comb sem limitação de colunas
    
    Tem_mais_prancha = False
    
    for i, row in dadosP2.iterrows():
        if row[0] == "Planta":
            c=i
            Tem_mais_prancha = True
            break

lista_titulos = lista_titulos.rename(columns={"Bitola (mm)": "Planta"})

lista_elem = pd.DataFrame()
# Cria lista com o nome do elemento e a linha em que ele aparece em relação ao DataFrame dad3tt
for i, row in dad3tt.iterrows():
    if row.iloc[0] != "50A" and row.iloc[0] !="60B":        
        elem = dad3tt.iloc[i].dropna().add(' ').sum()        
        elem1 = pd.DataFrame({"Elemento": [elem], "Linha": [i]})
        lista_elem = pd.concat([lista_elem, elem1]).reset_index(drop=True)

lista2=pd.DataFrame(columns= [0.0, 5.0, 6.3, 8.0, 10.0, 12.5, 16.0, 20.0, 25.0, 32.0])
lista2p=lista2
qtdA = lista2
it=0
while it < lista_elem.shape[0]-1:
    qtd = dad3tt.iloc[lista_elem.iloc[it,1]:lista_elem.iloc[it+1,1]].reset_index(drop=True).iloc[1:,:6]
    qtd = qtd.iloc[:,2:].astype(float).groupby(by=[2]).sum().reset_index().transpose().reset_index(drop=True)
    qtd.columns = qtd.iloc[0].reset_index(drop=True)
    qtd = qtd.iloc[1:].reset_index(drop=True)
    lista2 = pd.concat([lista2, preencher_nan(pd.concat([pd.DataFrame([lista_elem.iloc[it,0]]), qtd.iloc[2:]]))]).reset_index(drop=True)
    qtd = pd.concat([qtdA, qtd.iloc[2:]])
    lista2p = pd.concat([lista2p, preencher_nan(pd.concat([pd.DataFrame([lista_elem.iloc[it,0]]), qtd.mul([0., 154*1e-5, 245*1e-5, 395*1e-5, 617*1e-5, 963*1e-5, 1578*1e-5, 2466*1e-5, 3853*1e-5, 6313*1e-5], axis=1)]))])    
    it+=1    

qtd = dad3tt.iloc[lista_elem.iloc[it,1]:].reset_index(drop=True).iloc[1:,:6]
qtd = qtd.iloc[:,2:].astype(float).groupby(by=[2]).sum().reset_index().transpose().reset_index(drop=True)
qtd.columns = qtd.iloc[0].reset_index(drop=True)
qtd = qtd.iloc[1:].reset_index(drop=True)
lista2 = pd.concat([lista2, preencher_nan(pd.concat([pd.DataFrame([lista_elem.iloc[it,0]]), qtd.iloc[2:]]))]).dropna(how="all").reset_index(drop=True).reindex(sorted(lista2.columns), axis=1)
qtd = pd.concat([qtdA, qtd.iloc[2:]])
lista2p = pd.concat([lista2p, preencher_nan(pd.concat([pd.DataFrame([lista_elem.iloc[it,0]]), qtd.mul([0., 154*1e-5, 245*1e-5, 395*1e-5, 617*1e-5, 963*1e-5, 1578*1e-5, 2466*1e-5, 3853*1e-5, 6313*1e-5], axis=1)]))]).dropna(how="all").reset_index(drop=True)

lista2 = lista2.rename(columns={0.0: "Elemento"})
lista2p = lista2p.rename(columns={0.0: "Elemento"})



# Ordena o DataFrame por bitola e adiciona o nome da prancha
d2_comb = d2_comb.reindex(sorted(d2_comb.columns), axis=1)
d2_comb = pd.merge(lista_titulos.iloc[:,1], d2_comb, left_index=True, right_index=True)
d2_comb.loc["Total"] = d2_comb.sum(numeric_only=True)
d2_comb.at["Total", "Planta"] = "Total"


### Definição de interface

import customtkinter as ctk

root = ctk.CTk()
root.geometry("500x350")

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("dark-blue")

frame = ctk.CTkFrame(master=root)
frame.pack(pady=20, padx=60, fill="both", expand=True)

c1 = ctk.StringVar(value="on")
checkbox1 = ctk.CTkCheckBox(master=frame, text="Quantitativo por planta", variable=c1, onvalue="on", offvalue="off")
checkbox1.pack(pady=12, padx=20)

c2 = ctk.StringVar(value="on")
checkbox2 = ctk.CTkCheckBox(master=frame, text="Quantitativo por elemento", variable=c2, onvalue="on", offvalue="off")
checkbox2.pack(pady=12, padx=10)

c3 = ctk.StringVar(value="on")
checkbox3 = ctk.CTkCheckBox(master=frame, text="Comprimento (cm) por elemento", variable=c3, onvalue="on", offvalue="off")
checkbox3.pack(pady=12, padx=10)

def iniciar():
    # Exporta o DataFrame para excel 
    with pd.ExcelWriter("Quantitativo de aço.xlsx", engine="xlsxwriter") as writer:
        if c1.get() == "on":
            d2_comb.to_excel(writer, sheet_name="Quantitativo por planta", index = False)
        if c2.get() == "on":
            lista2p.to_excel(writer, sheet_name="Quantitativo por elemento", index = False)
        if c3.get() == "on":
            lista2.to_excel(writer, sheet_name="Comprimento (cm) por elemento", index = False)

button = ctk.CTkButton(master=frame, text="Iniciar", command=iniciar)
button.pack(pady=12, padx=10)

root.mainloop()

###

# Exporta o DataFrame para excel 
#with pd.ExcelWriter("Quantitativo de aço.xlsx", engine="xlsxwriter") as writer:
    #d2_comb.to_excel(writer, sheet_name="Quantitativo por planta", index = False)
    #lista2p.to_excel(writer, sheet_name="Quantitativo por elemento", index = False)
    #lista2.to_excel(writer, sheet_name="Comprimento (cm) por elemento", index = False)
        
### Define largura da coluna "Elemento"

from openpyxl import load_workbook

wb = load_workbook("Quantitativo de aço.xlsx")
if c2.get() == "on":
    ws = wb["Quantitativo por elemento"]
    
    for letter in ["A"]:
        max_width = 0
        
        for row_number in range (1, ws.max_row + 1):
            if len(ws[f'{letter}{row_number}'].value) > max_width:
                max_width = len(ws[f'{letter}{row_number}'].value)
                
        ws.column_dimensions[letter].width = max_width + 1
        
if c3.get() == "on":    
    ws = wb["Comprimento (cm) por elemento"]
    
    for letter in ["A"]:
        max_width = 0
        
        for row_number in range (1, ws.max_row + 1):
            if len(ws[f'{letter}{row_number}'].value) > max_width:
                max_width = len(ws[f'{letter}{row_number}'].value)
                
        ws.column_dimensions[letter].width = max_width + 1

wb.save("Quantitativo de aço.xlsx")

###


#with pd.ExcelWriter("teste.xlsx", engine="xlsxwriter", engine_kwargs={'options': {'strings_to_numbers': True}}) as writer:

      
#qtd.mul([0., 160*1e-5, 250*1e-5, 300*1e-5, 630*1e-5, 1000*1e-5, 1600*1e-5, 2500*1e-5, 4000*1e-5, 6300*1e-5], axis=1)])